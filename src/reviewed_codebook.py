"""Import reviewed keyword codebooks into runtime config structures."""

from __future__ import annotations

import csv
import datetime as dt
import hashlib
from pathlib import Path
from typing import Iterable

SOURCE_REQUIRED_COLUMNS = {
    "project_scope",
    "language",
    "source_layer",
    "code_family",
    "code",
    "label_en",
    "keyword",
    "reviewer",
    "review_decision",
    "suggested_replacement_keyword",
}

CSV_REQUIRED_COLUMNS = {
    "source_sheet",
    "source_row_id",
    "language",
    "code_family",
    "code",
    "label_en",
    "keyword_original",
    "reviewer",
    "review_decision",
    "keyword_final",
}

VALID_REVIEW_DECISIONS = {"no change", "fix", "delete"}
SUPPORTED_CODE_FAMILIES = {"friction", "topic", "sentiment"}


class ReviewedCodebookError(RuntimeError):
    """Raised when a reviewed codebook cannot be safely promoted."""


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _source_row_to_import_row(
    raw: dict[str, object],
    source_sheet: str,
    source_row_id: int | str,
    reviewed_at: str = "",
) -> dict[str, str]:
    decision = clean_text(raw.get("review_decision"))
    keyword_original = clean_text(raw.get("keyword"))
    replacement = clean_text(raw.get("suggested_replacement_keyword"))
    keyword_final = replacement
    if decision.lower() == "no change":
        keyword_final = keyword_original
    if decision.lower() == "delete":
        keyword_final = ""
    return {
        "source_sheet": source_sheet,
        "source_row_id": str(source_row_id),
        "project_scope": clean_text(raw.get("project_scope")),
        "language": clean_text(raw.get("language")),
        "source_layer": clean_text(raw.get("source_layer")),
        "code_family": clean_text(raw.get("code_family")),
        "code": clean_text(raw.get("code")),
        "label_en": clean_text(raw.get("label_en")),
        "keyword_original": keyword_original,
        "reviewer": clean_text(raw.get("reviewer")),
        "review_decision": decision,
        "keyword_final": keyword_final,
        "reviewed_at": reviewed_at,
    }


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise ReviewedCodebookError(f"Codebook CSV is empty: {path}")
        missing = sorted(CSV_REQUIRED_COLUMNS - set(reader.fieldnames))
        if missing:
            raise ReviewedCodebookError(f"Codebook CSV missing required columns: {missing}")
        return [{key: clean_text(value) for key, value in row.items()} for row in reader]


def _read_workbook_rows(path: Path, sheet_names: Iterable[str] | None = None) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise ReviewedCodebookError(
            "Reading .xlsx codebook sources requires openpyxl. "
            "Install project dependencies with `.venv/bin/pip install -r requirements.txt`."
        ) from error
    workbook = load_workbook(path, read_only=True, data_only=True)
    selected_sheets = list(sheet_names) if sheet_names else list(workbook.sheetnames)
    missing_sheets = sorted(set(selected_sheets) - set(workbook.sheetnames))
    if missing_sheets:
        raise ReviewedCodebookError(f"Workbook missing required sheets: {missing_sheets}")

    imported_rows: list[dict[str, str]] = []
    for sheet_name in selected_sheets:
        rows = list(workbook[sheet_name].iter_rows(values_only=True))
        if not rows:
            continue
        headers = [clean_text(value) for value in rows[0]]
        header_indexes = {name: index for index, name in enumerate(headers) if name}
        missing = sorted(SOURCE_REQUIRED_COLUMNS - set(header_indexes))
        if missing:
            raise ReviewedCodebookError(f"Workbook sheet {sheet_name!r} missing required columns: {missing}")
        for excel_row_number, values in enumerate(rows[1:], start=2):
            raw = {
                column: values[index] if index < len(values) else ""
                for column, index in header_indexes.items()
            }
            if not any(clean_text(raw.get(column)) for column in ("language", "code_family", "code", "keyword")):
                continue
            imported_rows.append(_source_row_to_import_row(raw, sheet_name, excel_row_number))
    return imported_rows


def load_reviewed_codebook_rows(
    path: Path,
    sheet_names: Iterable[str] | None = None,
) -> list[dict[str, str]]:
    if not path.exists():
        raise ReviewedCodebookError(f"Reviewed codebook source not found: {path}")
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _read_workbook_rows(path, sheet_names=sheet_names)
    if suffix == ".csv":
        return _read_csv_rows(path)
    raise ReviewedCodebookError(f"Unsupported codebook source type: {path.suffix}")


def _normal_language(value: str) -> str:
    return clean_text(value).lower()


def validate_reviewed_codebook_rows(
    rows: list[dict[str, str]],
    languages: Iterable[str],
    require_complete: bool = True,
) -> list[dict[str, str]]:
    requested = {_normal_language(language) for language in languages}
    selected = [
        row for row in rows
        if _normal_language(row.get("language", "")) in requested
        and clean_text(row.get("code_family", "")).lower() in SUPPORTED_CODE_FAMILIES
    ]
    present = {_normal_language(row.get("language", "")) for row in selected}
    missing_languages = sorted(requested - present)
    if missing_languages:
        raise ReviewedCodebookError(f"No reviewed codebook rows found for languages: {missing_languages}")

    blank_decisions = [row for row in selected if clean_text(row.get("review_decision", "")) == ""]
    if blank_decisions and require_complete:
        examples = [
            f"{row.get('source_sheet')}:{row.get('source_row_id')}"
            for row in blank_decisions[:8]
        ]
        raise ReviewedCodebookError(
            "Reviewed codebook has blank review_decision rows for requested language(s): "
            + ", ".join(examples)
        )

    invalid_decisions = [
        row for row in selected
        if clean_text(row.get("review_decision", "")).lower()
        and clean_text(row.get("review_decision", "")).lower() not in VALID_REVIEW_DECISIONS
    ]
    if invalid_decisions:
        examples = [
            f"{row.get('source_sheet')}:{row.get('source_row_id')}={row.get('review_decision')!r}"
            for row in invalid_decisions[:8]
        ]
        raise ReviewedCodebookError("Reviewed codebook has invalid review_decision rows: " + ", ".join(examples))

    if not require_complete:
        return selected

    fix_missing_replacement = [
        row for row in selected
        if clean_text(row.get("review_decision", "")).lower() == "fix"
        and clean_text(row.get("keyword_final", "")) == ""
    ]
    if fix_missing_replacement:
        examples = [
            f"{row.get('source_sheet')}:{row.get('source_row_id')}"
            for row in fix_missing_replacement[:8]
        ]
        raise ReviewedCodebookError("FIX rows missing keyword_final: " + ", ".join(examples))

    kept_missing_keyword = [
        row for row in selected
        if clean_text(row.get("review_decision", "")).lower() in {"no change", "fix"}
        and clean_text(row.get("keyword_final", "")) == ""
    ]
    if kept_missing_keyword:
        examples = [
            f"{row.get('source_sheet')}:{row.get('source_row_id')}"
            for row in kept_missing_keyword[:8]
        ]
        raise ReviewedCodebookError("Kept rows missing runtime keyword: " + ", ".join(examples))
    return selected


def build_runtime_config(
    rows: list[dict[str, str]],
    source_path: Path,
    languages: Iterable[str],
    command: str,
) -> dict:
    selected = validate_reviewed_codebook_rows(rows, languages=languages, require_complete=True)
    config = {
        "schema_version": "reviewed_codebook_runtime.v1",
        "generated_at": dt.datetime.now(dt.UTC).replace(microsecond=0).isoformat(),
        "command": command,
        "source": {
            "path": str(source_path),
            "sha256": sha256_file(source_path),
        },
        "languages": {},
    }
    for row in selected:
        language = clean_text(row["language"])
        family = clean_text(row["code_family"]).lower()
        code = clean_text(row["code"])
        if not code:
            raise ReviewedCodebookError(f"Reviewed codebook row missing code: {row.get('source_sheet')}:{row.get('source_row_id')}")
        language_config = config["languages"].setdefault(language, {"codes": {}})
        code_entry = language_config["codes"].setdefault(
            code,
            {
                "code_family": family,
                "label": clean_text(row.get("label_en")) or code,
                "keywords": [],
                "reviewed_rows": [],
            },
        )
        if code_entry["code_family"] != family:
            raise ReviewedCodebookError(f"Code has mixed code_family values: {language}/{code}")
        decision = clean_text(row.get("review_decision")).lower()
        keyword = clean_text(row.get("keyword_final"))
        if decision != "delete" and keyword and keyword not in code_entry["keywords"]:
            code_entry["keywords"].append(keyword)
        code_entry["reviewed_rows"].append({
            "source_sheet": clean_text(row.get("source_sheet")),
            "source_row_id": clean_text(row.get("source_row_id")),
            "review_decision": clean_text(row.get("review_decision")),
            "keyword_original": clean_text(row.get("keyword_original")),
            "keyword_final": keyword,
            "reviewer": clean_text(row.get("reviewer")),
            "reviewed_at": clean_text(row.get("reviewed_at")),
        })

    empty_codes = []
    for language, language_config in config["languages"].items():
        for code, code_entry in language_config["codes"].items():
            if not code_entry["keywords"]:
                empty_codes.append(f"{language}/{code}")
    if empty_codes:
        raise ReviewedCodebookError(f"Reviewed codebook has no kept keywords for codes: {empty_codes}")
    return config


def validation_status(rows: list[dict[str, str]], languages: Iterable[str]) -> dict:
    selected = validate_reviewed_codebook_rows(rows, languages=languages, require_complete=False)
    status: dict[str, dict] = {}
    for row in selected:
        language = clean_text(row.get("language"))
        language_status = status.setdefault(
            language,
            {
                "rows": 0,
                "blank_review_decision_rows": 0,
                "review_decision_counts": {},
                "code_family_counts": {},
            },
        )
        language_status["rows"] += 1
        decision = clean_text(row.get("review_decision"))
        if not decision:
            language_status["blank_review_decision_rows"] += 1
        else:
            language_status["review_decision_counts"][decision] = (
                language_status["review_decision_counts"].get(decision, 0) + 1
            )
        family = clean_text(row.get("code_family")).lower()
        language_status["code_family_counts"][family] = (
            language_status["code_family_counts"].get(family, 0) + 1
        )
    return status
