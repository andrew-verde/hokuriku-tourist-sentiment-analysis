#!/usr/bin/env python3
"""Export the reviewed Chinese codebook sheet from an Excel workbook to a UTF-8 CSV template.

This script extracts the Chinese-language rows from the multilingual keyword codebook review,
validates reviewer decisions, and outputs a canonical CSV that is imported by the social-media pipeline.
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook

# This module reads the Excel workbook of reviewed keyword codes, extracts the Chinese rows,
# and exports them as a CSV template with columns for code families, keywords, and review decisions.

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_WORKBOOK = ROOT / "docs" / "codebook_reviews" / "source" / "multilingual_keyword_codebook_review.xlsx"
DEFAULT_OUTPUT = ROOT / "docs" / "codebook_templates" / "chinese_reviewed_codebook_template.csv"
CHINESE_SHEET = "Chinese"

OUTPUT_COLUMNS = [
    "source_sheet",
    "source_row_id",
    "project_scope",
    "language",
    "source_layer",
    "code_family",
    "code",
    "label_en",
    "label_cn",
    "keyword_original",
    "keyword_translation_or_note",
    "current_pipeline_status",
    "reviewer",
    "review_decision",
    "keyword_final",
    "notes",
    "reviewed_at",
]

SOURCE_TO_OUTPUT = {
    "project_scope": "project_scope",
    "language": "language",
    "source_layer": "source_layer",
    "code_family": "code_family",
    "code": "code",
    "label_en": "label_en",
    "label_cn": "label_cn",
    "keyword": "keyword_original",
    "keyword_translation_or_note": "keyword_translation_or_note",
    "current_pipeline_status": "current_pipeline_status",
    "reviewer": "reviewer",
    "review_decision": "review_decision",
    "suggested_replacement_keyword": "keyword_final",
    "notes": "notes",
}

VALID_DECISIONS = {"No change", "FIX", "delete"}
PROJECT_SCOPE_RENAMES = {
    "cross_language_tourism_group_project": "cross_language_tourism_group_project",
}


class CodebookExportError(RuntimeError):
    pass


def _clean(value: object) -> str:
    """Normalize Excel cell values to plain text: convert None to empty string and strip whitespace."""
    return "" if value is None else str(value).strip()


def export_chinese_codebook_template(
    workbook_path: Path = DEFAULT_WORKBOOK,
    output_path: Path = DEFAULT_OUTPUT,
    reviewed_at: str = "",
) -> list[dict[str, str]]:
    """Extract Chinese rows from the reviewed codebook Excel workbook and write to CSV."""
    # Read the Excel workbook and extract the Chinese sheet.
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if CHINESE_SHEET not in workbook.sheetnames:
        raise CodebookExportError(f"Workbook missing required sheet: {CHINESE_SHEET}")

    rows = list(workbook[CHINESE_SHEET].iter_rows(values_only=True))
    if not rows:
        raise CodebookExportError("Chinese sheet is empty")

    # Parse header row to find column indexes.
    headers = [_clean(value) for value in rows[0]]
    header_indexes = {name: index for index, name in enumerate(headers) if name}
    missing = sorted(set(SOURCE_TO_OUTPUT) - set(header_indexes))
    if missing:
        raise CodebookExportError(f"Chinese sheet missing required columns: {missing}")

    exported: list[dict[str, str]] = []
    for excel_row_number, source_row in enumerate(rows[1:], start=2):
        # Start with empty row, map columns from SOURCE_TO_OUTPUT, validate review decision.
        row = {column: "" for column in OUTPUT_COLUMNS}
        row["source_sheet"] = CHINESE_SHEET
        row["source_row_id"] = str(excel_row_number)
        row["reviewed_at"] = reviewed_at
        # Map source Excel columns to output CSV columns.
        for source_col, output_col in SOURCE_TO_OUTPUT.items():
            source_index = header_indexes[source_col]
            row[output_col] = _clean(source_row[source_index] if source_index < len(source_row) else "")
        row["project_scope"] = PROJECT_SCOPE_RENAMES.get(row["project_scope"], row["project_scope"])

        # Skip rows with missing core fields or non-Chinese language.
        if not any(row[column] for column in ("language", "code_family", "code", "keyword_original")):
            continue
        if row["language"] != "Chinese":
            continue
        # Validate that review_decision is one of {No change, FIX, delete}.
        if row["review_decision"] not in VALID_DECISIONS:
            raise CodebookExportError(
                f"Invalid review_decision on Chinese row {excel_row_number}: {row['review_decision']!r}"
            )
        # Set keyword_final based on the review decision.
        if row["review_decision"] == "No change":
            row["keyword_final"] = row["keyword_original"]
        if row["review_decision"] == "delete":
            row["keyword_final"] = ""
        if row["review_decision"] == "FIX" and not row["keyword_final"]:
            raise CodebookExportError(f"FIX row missing suggested replacement on Chinese row {excel_row_number}")
        exported.append(row)

    # Write the exported rows to a CSV file with UTF-8 signature.
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(exported)
    return exported


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments for workbook path, output path, and review timestamp."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--reviewed-at", default="")
    return parser.parse_args()


def main() -> int:
    """Orchestrate: load workbook, extract Chinese rows, export to CSV, and report row count."""
    args = parse_args()
    try:
        rows = export_chinese_codebook_template(args.workbook, args.output, args.reviewed_at)
    except CodebookExportError as error:
        print(error)
        return 1
    print(f"Exported {len(rows)} Chinese codebook rows to {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
